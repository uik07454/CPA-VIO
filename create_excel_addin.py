"""
Create Excel Add-in with Checkboxes and Delete Button

This creates an Excel file with:
- Real checkboxes (not dropdown)
- "Delete Selected Rows" button
- No need to save/close/reopen

Usage:
    python create_excel_addin.py <source_excel_file>
"""

import openpyxl
from pathlib import Path
from openpyxl.styles import Font, PatternFill, Alignment


# VBA code for the Excel workbook
VBA_CODE = '''
' Module: DeleteToggleModule

Sub DeleteSelectedRows()
    Dim ws As Worksheet
    Dim lastRow As Long
    Dim i As Long
    Dim deleteCount As Integer
    Dim chk As CheckBox
    Dim checkboxName As String
    Dim rowNum As Long
    Dim response As VbMsgBoxResult
    
    Set ws = ActiveSheet
    deleteCount = 0
    
    ' Disable screen updating for better performance
    Application.ScreenUpdating = False
    Application.EnableEvents = False
    
    On Error GoTo ErrorHandler
    
    ' Count how many checkboxes are checked
    For Each chk In ws.CheckBoxes
        If chk.Value = xlOn Then
            deleteCount = deleteCount + 1
        End If
    Next chk
    
    If deleteCount = 0 Then
        MsgBox "No rows selected for deletion.", vbInformation, "Delete Rows"
        GoTo Cleanup
    End If
    
    ' Confirm deletion
    response = MsgBox("Delete " & deleteCount & " selected row(s)?", _
                      vbYesNo + vbQuestion, "Confirm Deletion")
    
    If response <> vbYes Then
        GoTo Cleanup
    End If
    
    ' Delete rows in reverse order (bottom to top) to avoid index shifting
    lastRow = ws.Cells(ws.Rows.Count, 1).End(xlUp).Row
    
    For i = lastRow To 2 Step -1
        checkboxName = "chk_row_" & i
        
        On Error Resume Next
        Set chk = ws.CheckBoxes(checkboxName)
        On Error GoTo ErrorHandler
        
        If Not chk Is Nothing Then
            If chk.Value = xlOn Then
                ' Delete the entire row
                ws.Rows(i).Delete
            End If
        End If
        Set chk = Nothing
    Next i
    
    MsgBox "Deleted " & deleteCount & " row(s) successfully!", vbInformation, "Complete"
    
Cleanup:
    Application.ScreenUpdating = True
    Application.EnableEvents = True
    Exit Sub
    
ErrorHandler:
    Application.ScreenUpdating = True
    Application.EnableEvents = True
    MsgBox "Error: " & Err.Description, vbCritical, "Error"
End Sub
'''


def create_excel_with_checkboxes(source_file: str, output_file: str = None):
    """
    Create Excel file with checkboxes and delete button.
    
    Note: openpyxl cannot create form controls (checkboxes, buttons).
    This function creates the Excel structure and provides VBA code.
    You must add the checkboxes and button manually in Excel.
    
    Args:
        source_file: Source Excel file path
        output_file: Output file path (default: source with _with_checkboxes suffix)
    """
    if output_file is None:
        source_path = Path(source_file)
        output_file = source_path.parent / f"{source_path.stem}_with_checkboxes.xlsm"
    
    # Load workbook
    wb = openpyxl.load_workbook(source_file)
    ws = wb.active
    
    # Add a column for checkboxes (column A, shift everything right)
    ws.insert_cols(1)
    
    # Set header
    ws['A1'] = "Select"
    ws['A1'].font = Font(bold=True, color="FFFFFF")
    ws['A1'].alignment = Alignment(horizontal="center", vertical="center")
    ws['A1'].fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    
    # Set column width for checkboxes
    ws.column_dimensions['A'].width = 8
    
    # Save as macro-enabled workbook
    wb.save(output_file)
    wb.close()
    
    # Create instruction file
    instruction_file = Path(output_file).with_suffix('.instructions.txt')
    with open(instruction_file, 'w') as f:
        f.write("=" * 70 + "\n")
        f.write("EXCEL CHECKBOX SETUP INSTRUCTIONS\n")
        f.write("=" * 70 + "\n\n")
        f.write("Step 1: Open the file in Excel\n")
        f.write(f"  File: {output_file}\n\n")
        
        f.write("Step 2: Enable Developer Tab\n")
        f.write("  1. File → Options → Customize Ribbon\n")
        f.write("  2. Check 'Developer' on the right side\n")
        f.write("  3. Click OK\n\n")
        
        f.write("Step 3: Add Checkboxes\n")
        f.write("  1. Click 'Developer' tab\n")
        f.write("  2. Click 'Insert' → Form Controls → Checkbox\n")
        f.write("  3. Draw a small checkbox in cell A2\n")
        f.write("  4. Right-click checkbox → Format Control\n")
        f.write("  5. Set 'Cell link' to a hidden column (e.g., Z2)\n")
        f.write("  6. Delete the checkbox text label\n")
        f.write("  7. Resize checkbox to fit in cell\n")
        f.write("  8. Copy checkbox and paste to A3, A4, A5... for all rows\n\n")
        
        f.write("Step 4: Add Delete Button\n")
        f.write("  1. Developer tab → Insert → Form Controls → Button\n")
        f.write("  2. Draw button at top of sheet (e.g., above column A)\n")
        f.write("  3. When 'Assign Macro' dialog appears, click 'New'\n")
        f.write("  4. Delete the auto-generated code\n")
        f.write("  5. Copy the VBA code from the .vba.txt file\n")
        f.write("  6. Paste it into the VBA editor\n")
        f.write("  7. Close VBA editor\n")
        f.write("  8. Change button text to 'Delete Selected Rows'\n\n")
        
        f.write("Step 5: Use the Feature\n")
        f.write("  1. Check boxes for rows you want to delete\n")
        f.write("  2. Click 'Delete Selected Rows' button\n")
        f.write("  3. Confirm deletion\n")
        f.write("  4. Done! Rows are deleted instantly\n\n")
        
        f.write("=" * 70 + "\n")
        f.write("ALTERNATIVE: Use the Auto-Delete Watcher (Much Easier!)\n")
        f.write("=" * 70 + "\n\n")
        f.write("Instead of manual checkbox setup, use:\n")
        f.write("  python auto_delete_watcher.py output/changes.xlsx\n\n")
        f.write("Then just mark rows TRUE and save - auto-deletes!\n")
        f.write("No checkbox setup needed!\n")
    
    # Save VBA code
    vba_file = Path(output_file).with_suffix('.vba.txt')
    with open(vba_file, 'w') as f:
        f.write("=" * 70 + "\n")
        f.write("VBA CODE - Copy to Excel VBA Editor\n")
        f.write("=" * 70 + "\n\n")
        f.write(VBA_CODE)
    
    print(f"✓ Created: {output_file}")
    print(f"✓ Instructions: {instruction_file}")
    print(f"✓ VBA Code: {vba_file}")
    print()
    print("⚠️  WARNING: Adding checkboxes manually is VERY tedious!")
    print("    You need to add one checkbox for EACH row.")
    print()
    print("💡 RECOMMENDED: Use auto_delete_watcher.py instead!")
    print("    python auto_delete_watcher.py output/changes.xlsx")
    print("    Much easier - just mark TRUE and save!")


if __name__ == "__main__":
    import sys
    
    if len(sys.argv) < 2:
        print("=" * 70)
        print("CREATE EXCEL WITH CHECKBOXES AND DELETE BUTTON")
        print("=" * 70)
        print()
        print("Usage:")
        print("  python create_excel_addin.py <excel_file>")
        print()
        print("⚠️  WARNING: This requires MANUAL checkbox setup in Excel!")
        print("   You must add a checkbox for EVERY row - very tedious!")
        print()
        print("💡 BETTER OPTION: Use auto_delete_watcher.py instead")
        print("   python auto_delete_watcher.py <excel_file>")
        print("   Just mark rows TRUE and save - auto-deletes!")
        print("=" * 70)
        sys.exit(1)
    
    source_file = sys.argv[1]
    create_excel_with_checkboxes(source_file)
