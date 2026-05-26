# output.py
"""
Output rendering utilities.

Provides functions to render a list of ``ChangeRecord`` objects as an Excel
(.xlsx) workbook with hyperlinks into the annotated .docx.
"""

import difflib
import shutil
from io import BytesIO
from pathlib import Path
from copy import copy

import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.cell.rich_text import CellRichText, TextBlock, InlineFont
from openpyxl.drawing.image import Image as XlImage
from openpyxl.drawing.spreadsheet_drawing import TwoCellAnchor, AnchorMarker, AnchorClientData
from openpyxl.utils import column_index_from_string, coordinate_to_tuple
from openpyxl.utils.units import pixels_to_EMU

from run_utils import resolve_color
from constants import (
    ChangeType, DEFAULT_TEMPLATE_SHEET, EMPTY_FIELD, make_bookmark_id,
    IMAGE_WIDTH_PX, IMAGE_HEIGHT_PX, IMAGE_SIZE_MODE,
    IMAGE_ROW_HEIGHT_MAX, IMAGE_ROW_HEIGHT_MIN, IMAGE_SIZE_MODE_ORIGINAL,
    DIFF_DEL_COLOR, DIFF_INS_COLOR, DIFF_EQ_COLOR,
    PARA_DELIMITER, PARA_OUTPUT_DELIMITER,
    CELL_DELIMITER, CELL_OUTPUT_DELIMITER,
    LINE_BREAK_DELIMITER, LINE_BREAK_OUTPUT_DELIMITER,
    IMAGE_PX_MIN, IMAGE_PX_MAX,
)

# AppConfig is imported lazily to avoid a hard dependency when running CLI-only.
from typing import TYPE_CHECKING
if TYPE_CHECKING:
    from app_config import AppConfig


def _tokenise(text: str) -> list[str]:
    """
    Split *text* into a word token list, collapsing consecutive spaces/tabs.

    Each line is tokenised independently; a ``"\\n"`` sentinel is inserted
    between lines so the diff can align line breaks correctly.
    Consecutive whitespace within a line is collapsed to a single space.
    """
    tokens: list[str] = []
    for line_idx, line in enumerate(text.split("\n")):
        if line_idx > 0:
            tokens.append("\n")
        tokens.extend(word for word in line.split() if word)
    return tokens


def _runs_from_tokens(
    tokens: list[str], 
    color: str, 
    is_last_opcode: bool
) -> list[TextBlock]:
    """
    Build ``TextBlock`` runs from *tokens*.

    ``"\\n"`` tokens are embedded into the preceding text run (e.g. ``"word\\n"``)
    so that the ``<t>`` element always contains non-whitespace content, which
    causes openpyxl's ``whitespace()`` to add ``xml:space="preserve"`` — required
    for Excel to preserve the newline without triggering a repair warning.
    A trailing space is appended to the last run when *is_last_opcode* is False.
    """
    runs: list[TextBlock] = []
    font = InlineFont(rFont="Consolas", color=color)
    buf: list[str] = []
    pending_newline = False

    def _flush(trailing: str = "") -> None:
        if not buf:
            return
        text = " ".join(buf) + trailing
        runs.append(TextBlock(font, text))
        buf.clear()

    for tok in tokens:
        if tok == "\n":
            # Attach the newline to the end of the current word buffer so the
            # <t> element is non-empty, guaranteeing xml:space="preserve".
            if buf:
                _flush(trailing="\n")
            else:
                # No preceding words in this segment — carry the newline forward.
                pending_newline = True
        else:
            if pending_newline:
                buf.append("\n" + tok)
                pending_newline = False
            else:
                buf.append(tok)

    _flush(trailing=" " if not is_last_opcode else "")
    return runs


def _make_diff_rich_text(
    old_text: str,
    new_text: str,
    del_color: str,
    ins_color: str,
    eq_color: str,
) -> tuple[CellRichText, CellRichText]:
    """
    Compute a word-level diff and return ``(old_rich, new_rich)`` CellRichText objects.

    Deleted words are coloured red in *old_rich*; inserted words are coloured
    green in *new_rich*; equal words are black in both.  Newlines are preserved
    as explicit line-break runs so Excel renders them correctly.
    """
    old_tokens = _tokenise(old_text)
    new_tokens = _tokenise(new_text)

    matcher = difflib.SequenceMatcher(None, old_tokens, new_tokens, autojunk=False)

    old_runs: list[TextBlock] = []
    new_runs: list[TextBlock] = []

    opcodes = matcher.get_opcodes()
    for idx, (tag, i1, i2, j1, j2) in enumerate(opcodes):
        is_last = idx == len(opcodes) - 1

        if tag == "equal":
            runs = _runs_from_tokens(old_tokens[i1:i2], eq_color, is_last)
            old_runs.extend(runs)
            new_runs.extend(_runs_from_tokens(old_tokens[i1:i2], eq_color, is_last))
        elif tag == "replace":
            old_runs.extend(_runs_from_tokens(old_tokens[i1:i2], del_color, is_last))
            new_runs.extend(_runs_from_tokens(new_tokens[j1:j2], ins_color, is_last))
        elif tag == "delete":
            old_runs.extend(_runs_from_tokens(old_tokens[i1:i2], del_color, is_last))
        elif tag == "insert":
            new_runs.extend(_runs_from_tokens(new_tokens[j1:j2], ins_color, is_last))

    # Guard: CellRichText must not be empty — fall back to plain black text
    if not old_runs:
        old_runs = [TextBlock(InlineFont(rFont="Consolas", color=eq_color), old_text or " ")]
    if not new_runs:
        new_runs = [TextBlock(InlineFont(rFont="Consolas", color=eq_color), new_text or " ")]

    return CellRichText(*old_runs), CellRichText(*new_runs)


def _embed_image(
    ws,
    image_bytes: bytes,
    cell_address: str,
    width_px: int,
    height_px: int,
) -> None:
    """Embed *image_bytes* into *ws* anchored at *cell_address*.

    Uses TwoCellAnchor with fLocksWithSheet so the image hides with its row
    when rows are filtered, instead of floating freely over other rows.
    """
    img = XlImage(BytesIO(image_bytes))
    img.width  = width_px
    img.height = height_px

    row, col = coordinate_to_tuple(cell_address.upper())  # 1-based

    # _from: top-left corner of the target cell (zero offset).
    # to: bottom-right extent of the image within the same cell (in EMU).
    from_marker = AnchorMarker(col=col - 1, colOff=0, row=row - 1, rowOff=0)
    to_marker   = AnchorMarker(col=col - 1, colOff=pixels_to_EMU(width_px),
                               row=row - 1, rowOff=pixels_to_EMU(height_px))

    anchor            = TwoCellAnchor(editAs="twoCell")
    anchor._from      = from_marker
    anchor.to         = to_marker
    # fLocksWithSheet=True causes Excel to hide the image when its row is
    # filtered out, preventing images from floating over unrelated rows.
    anchor.clientData = AnchorClientData(fLocksWithSheet=True)

    img.anchor = anchor
    ws.add_image(img)


def _calculate_image_dimensions(
    record_image_size: tuple[int, int] | None,
    default_width_px: int,
    default_height_px: int,
    mode: str = IMAGE_SIZE_MODE,
) -> tuple[int, int]:
    """
    Calculate image dimensions for embedding.

    In 'fixed' mode, always returns the configured defaults.
    In 'original' mode, uses the record's wp:extent size (clamped) when available,
    falling back to the configured defaults.
    """
    if mode == IMAGE_SIZE_MODE_ORIGINAL and record_image_size is not None:
        width, height = record_image_size
        width  = max(IMAGE_PX_MIN, min(IMAGE_PX_MAX, width))
        height = max(IMAGE_PX_MIN, min(IMAGE_PX_MAX, height))
        return width, height
    return default_width_px, default_height_px


def _calculate_row_height(image_height_px: int, default_image_height_px: int) -> int:
    """
    Calculate row height in points based on the image height in pixels.

    The default row height is derived from the default image height, so the row height
    always scales from the configured image size rather than a separate row-height setting.
    """
    # Excel row height is approximately 0.75 times the pixel height, plus a small padding.
    default_row_height_pt = int(default_image_height_px * 0.75) + 20
    calculated_height_pt = int(image_height_px * 0.75) + 20
    if calculated_height_pt >= IMAGE_ROW_HEIGHT_MIN and calculated_height_pt <= IMAGE_ROW_HEIGHT_MAX:
        return calculated_height_pt
    return default_row_height_pt


def _format_content(text: str, empty_field: str) -> str:
    """Replace internal delimiters and strip whitespace for Excel display."""
    return (
        (text or empty_field)
        .replace(CELL_DELIMITER, CELL_OUTPUT_DELIMITER)
        .replace(PARA_DELIMITER, PARA_OUTPUT_DELIMITER)
        .replace(LINE_BREAK_DELIMITER, LINE_BREAK_OUTPUT_DELIMITER)
        .strip()
    )


def _build_heading_cell_value(r, abs_annotated: str) -> str:
    """Return a HYPERLINK formula string if *r* has a bookmark, otherwise the plain heading."""
    if r.bookmark_id is None:
        return r.heading
    bookmark_id = make_bookmark_id(r.bookmark_id)
    url = f"{abs_annotated}#{bookmark_id}"
    return f'=HYPERLINK("{url}", "{r.heading}")'


# col_idx: (col_heading, col_change_type, col_old, col_new) — 1-based column indices.
_ColIdx = tuple[int, int, int, int]

# Default column indices matching the scratch generator's fixed layout (A=1, B=2, C=3, D=4).
_DEFAULT_COL_IDX: _ColIdx = (1, 2, 3, 4)


def _col_idx_from_letters(columns: list[str]) -> _ColIdx:
    """Convert a list of 4 Excel column letters to a tuple of 1-based column indices."""
    return tuple(column_index_from_string(c) for c in columns)


def _apply_color_style(
    ws, row_idx: int, role: str,
    old_color: str | None, new_color: str | None,
    col_old: int, col_new: int,
) -> None:
    """Apply font-colour or highlight-colour styling to the old/new content cells."""
    cell_old = ws.cell(row=row_idx, column=col_old)
    cell_new = ws.cell(row=row_idx, column=col_new)
    if role == "font":
        cell_old.font = cell_old.font.copy(name="Consolas", color=old_color or "000000")
        cell_new.font = cell_new.font.copy(name="Consolas", color=new_color or "000000")
        return
    # Highlight colour change — colour the cell background
    cell_old.font = cell_old.font.copy(name="Consolas")
    cell_new.font = cell_new.font.copy(name="Consolas")
    if old_color:
        cell_old.fill = PatternFill("solid", fgColor=old_color)
    if new_color:
        cell_new.fill = PatternFill("solid", fgColor=new_color)


def _apply_content_style(
    ws, row_idx: int, r, old_text: str, new_text: str,
    del_color: str, ins_color: str, eq_color: str,
    col_old: int, col_new: int,
) -> None:
    """Apply inline diff, strikethrough, or default monospace styling to content cells."""
    if r.change_type == ChangeType.MODIFY_CONTENT:
        old_rich, new_rich = _make_diff_rich_text(old_text, new_text, del_color, ins_color, eq_color)
        ws.cell(row=row_idx, column=col_old).value = old_rich
        ws.cell(row=row_idx, column=col_new).value = new_rich
        return

    if r.change_type in (ChangeType.STRIKETHROUGH_ADDED, ChangeType.STRIKETHROUGH_REMOVED):
        strike_added = r.change_type == ChangeType.STRIKETHROUGH_ADDED
        cell_old = ws.cell(row=row_idx, column=col_old)
        cell_new = ws.cell(row=row_idx, column=col_new)
        cell_old.font = cell_old.font.copy(name="Consolas", strike=not strike_added)
        cell_new.font = cell_new.font.copy(name="Consolas", strike=strike_added)
        return

    cell_old = ws.cell(row=row_idx, column=col_old)
    cell_new = ws.cell(row=row_idx, column=col_new)
    cell_old.font = cell_old.font.copy(name="Consolas")
    cell_new.font = cell_new.font.copy(name="Consolas")


def _apply_row_style(
    ws, row_idx: int, r, old_text: str, new_text: str,
    del_color: str, ins_color: str, eq_color: str,
    col_idx: _ColIdx = _DEFAULT_COL_IDX,
) -> None:
    """Dispatch styling for a record row: colour-role styles take priority over content styles."""
    _, _, col_old, col_new = col_idx
    role      = r.style_meta.get("color_role")
    old_color = resolve_color(r.style_meta.get("old_color"))
    new_color = resolve_color(r.style_meta.get("new_color"))

    if role in ("font", "highlight"):
        _apply_color_style(ws, row_idx, role, old_color, new_color, col_old, col_new)
        return

    _apply_content_style(ws, row_idx, r, old_text, new_text, del_color, ins_color, eq_color, col_old, col_new)


def _embed_row_images(
    ws, row_idx: int, r,
    width_px: int, height_px: int, default_image_height_px: int,
    mode: str = IMAGE_SIZE_MODE,
    col_idx: _ColIdx = _DEFAULT_COL_IDX,
) -> None:
    """Embed old/new images into the content cells if present, clearing the text value."""
    _, _, col_old, col_new = col_idx
    actual_row_height_pt = []
    if r.old_image_data is not None:
        col_letter = ws.cell(row=row_idx, column=col_old).column_letter
        img_w, img_h = _calculate_image_dimensions(r.old_image_size_px, width_px, height_px, mode)
        actual_row_height_pt.append(_calculate_row_height(img_h, default_image_height_px))
        _embed_image(ws, r.old_image_data, f"{col_letter}{row_idx}", img_w, img_h)
        ws.cell(row=row_idx, column=col_old).value = None  # clear text; image floats over cell

    if r.new_image_data is not None:
        col_letter = ws.cell(row=row_idx, column=col_new).column_letter
        img_w, img_h = _calculate_image_dimensions(r.new_image_size_px, width_px, height_px, mode)
        actual_row_height_pt.append(_calculate_row_height(img_h, default_image_height_px))
        _embed_image(ws, r.new_image_data, f"{col_letter}{row_idx}", img_w, img_h)
        ws.cell(row=row_idx, column=col_new).value = None  # clear text; image floats over cell

    if actual_row_height_pt:
        ws.row_dimensions[row_idx].height = max(actual_row_height_pt)


def generate_excel_output(
    records,
    annotated_docx_path: str,
    out_path: str,
    config: "AppConfig | None" = None,
) -> None:
    """Generate an Excel (.xlsx) change table with hyperlinks into the annotated .docx.

    Each heading cell uses Excel's ``HYPERLINK()`` formula to jump to the exact
    paragraph in the annotated document via a ``#bookmark`` fragment.

    Args:
        records: An iterable of ``ChangeRecord`` objects.
        annotated_docx_path: Path to the annotated ``.docx`` file that contains
            the injected bookmarks.
        out_path: Destination path for the ``.xlsx`` output file.
        config: Optional ``AppConfig`` instance; falls back to module-level constants when None.
    """
    # Resolve effective rendering values from config or module-level constants.
    empty_field          = config.empty_field         if config else EMPTY_FIELD
    img_w                = config.image_width_px      if config else IMAGE_WIDTH_PX
    img_h                = config.image_height_px     if config else IMAGE_HEIGHT_PX
    default_image_height = config.image_height_px     if config else IMAGE_HEIGHT_PX
    img_mode             = config.image_size_mode     if config else IMAGE_SIZE_MODE
    del_color            = config.diff_del_color      if config else DIFF_DEL_COLOR
    ins_color            = config.diff_ins_color      if config else DIFF_INS_COLOR
    eq_color             = config.diff_eq_color       if config else DIFF_EQ_COLOR

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Changes"

    # Header row
    headers = ["Heading / Sub-heading", "Change Type", "Old Content", "New Content"]
    ws.append(headers)

    abs_annotated = Path(annotated_docx_path).name

    # Hyperlink font style (blue + underline, like a real hyperlink)
    hyperlink_font = Font(color="0563C1", underline="single")

    for r in records:
        row_idx = ws.max_row + 1
        old_text = _format_content(r.old_text, empty_field)
        new_text = _format_content(r.new_text, empty_field)

        ws.append([_build_heading_cell_value(r, abs_annotated), r.change_type, old_text, new_text])

        # Style the heading cell to look like a hyperlink
        if r.bookmark_id is not None:
            ws.cell(row=row_idx, column=1).font = hyperlink_font

        # Apply colour styling based on style_meta
        _apply_row_style(ws, row_idx, r, old_text, new_text, del_color, ins_color, eq_color)
        _embed_row_images(ws, row_idx, r, img_w, img_h, default_image_height, img_mode)

        # Prevent Excel from treating formula-like text as a formula.
        for col in (3, 4):
            cell = ws.cell(row=row_idx, column=col)
            if cell.data_type == "f":
                cell.data_type = "s"        # change data type to literal string
                cell.quotePrefix = True     # add single quote prefix, e.g. ('=0)

    wb.save(out_path)
    print(f"[INFO] Excel output saved to: {out_path}")


def generate_excel_from_template(
    records,
    annotated_docx_path: str,
    out_path: str,
    template_path: str,
    column_mapping: list[str],
    config: "AppConfig | None" = None,
    sheet_name: str = DEFAULT_TEMPLATE_SHEET,
) -> None:
    """Copy *template_path* to *out_path* and write change records into the target sheet.

    The template's header row (row 1) and all formatting are preserved.
    Only the cells written by this function are modified.

    Args:
        records: An iterable of ``ChangeRecord`` objects.
        annotated_docx_path: Path to the annotated ``.docx`` file.
        out_path: Destination path for the output ``.xlsx`` file.
        template_path: Path to the source Excel template file.
        column_mapping: List of 4 Excel column letters [heading, change_type, old, new].
        config: Optional ``AppConfig`` instance; falls back to module-level constants when None.
        sheet_name: Target sheet name; falls back to the first sheet if not found.
    """
    empty_field          = config.empty_field     if config else EMPTY_FIELD
    img_w                = config.image_width_px  if config else IMAGE_WIDTH_PX
    img_h                = config.image_height_px if config else IMAGE_HEIGHT_PX
    default_image_height = config.image_height_px if config else IMAGE_HEIGHT_PX
    img_mode             = config.image_size_mode if config else IMAGE_SIZE_MODE
    del_color            = config.diff_del_color  if config else DIFF_DEL_COLOR
    ins_color            = config.diff_ins_color  if config else DIFF_INS_COLOR
    eq_color             = config.diff_eq_color   if config else DIFF_EQ_COLOR

    # Copy the template — the original must never be modified.
    shutil.copy2(template_path, out_path)

    wb = openpyxl.load_workbook(out_path)
    ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.worksheets[0]

    # Clear fixed heights on all data rows so Excel auto-fits on open.
    for rd in ws.row_dimensions.values():
        if rd.index >= 2:
            rd.height = None

    col_idx: _ColIdx = _col_idx_from_letters(column_mapping)
    col_heading, col_change_type, col_old, col_new = col_idx

    abs_annotated  = Path(annotated_docx_path).name

    # Start writing from row 2; row 1 is the template header — never touch it.
    next_row = 2
    for r in records:
        row_idx  = next_row
        next_row += 1

        old_text = _format_content(r.old_text, empty_field)
        new_text = _format_content(r.new_text, empty_field)

        ws.cell(row=row_idx, column=col_heading).value     = _build_heading_cell_value(r, abs_annotated)
        ws.cell(row=row_idx, column=col_change_type).value = r.change_type
        ws.cell(row=row_idx, column=col_old).value         = old_text
        ws.cell(row=row_idx, column=col_new).value         = new_text

        # Prevent Excel from treating formula-like text as a formula.
        for col in (col_old, col_new):
            cell = ws.cell(row=row_idx, column=col)
            if cell.data_type == "f":
                cell.data_type = "s"
                cell.quotePrefix = True

        if r.bookmark_id is not None:
            heading_cell = ws.cell(row=row_idx, column=col_heading)
            hc = copy(heading_cell.font)
            hc.color = "0563C1"
            hc.underline = "single"
            heading_cell.font = hc

        _apply_row_style(ws, row_idx, r, old_text, new_text, del_color, ins_color, eq_color, col_idx)
        _embed_row_images(ws, row_idx, r, img_w, img_h, default_image_height, img_mode, col_idx)

    wb.save(out_path)
    print(f"[INFO] Template-based Excel output saved to: {out_path}")
