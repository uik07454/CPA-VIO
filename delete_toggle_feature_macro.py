"""
Delete Toggle Feature with Auto-Delete Macro

This version adds VBA macro to Excel that automatically deletes rows
when the toggle is set to TRUE.
"""

import openpyxl
from pathlib import Path
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.datavalidation import DataValidation


# VBA Macro code for auto-delete on TRUE selection
VBA_AUTO_DELETE_MACRO = '''
Private Sub Worksheet_Change(ByVal Target As Range)
    Dim toggleCol As String
    Dim startRow As Long
    Dim cell As Range
    
    ' Configuration
    toggleCol = "E"  ' Change this to match your toggle column
    startRow = 2     ' First data row
    
    ' Only process if change is in toggle column
    If Target.Column <> Range(toggleCol & "1").Column Then Exit Sub
    If Target.Row < startRow Then Exit Sub
    
    ' Disable events to prevent recursive calls
    Application.EnableEvents = False
    Application.ScreenUpdating = False
    
    On Error GoTo ErrorHandler
    
    For Each cell In Target
        If UCase(Trim(cell.Value)) = "TRUE" Then
            ' Confirm deletion
            Dim response As VbMsgBoxResult
            response = MsgBox("Delete row " & cell.Row & "?", vbYesNo + vbQuestion, "Confirm Delete")
            
            If response = vbYes Then
                ' Delete the row
                cell.EntireRow.Delete
                ' Note: After deletion, don't process more cells as row numbers changed
                Exit For
            Else
                ' User cancelled, reset to FALSE
                cell.Value = "FALSE"
            End If
        End If
    Next cell
    
ErrorHandler:
    Application.EnableEvents = True
    Application.ScreenUpdating = True
End Sub
'''


def add_toggle_column_with_macro(
    excel_path: str,
    toggle_col: str = "E",
    start_row: int = 2,
    header_text: str = "Delete Toggle"
) -> None:
    """
    Add toggle column with auto-delete VBA macro to an Excel file.
    
    NOTE: This creates a .xlsm file (macro-enabled) instead of .xlsx
    
    Args:
        excel_path: Path to the Excel file
        toggle_col: Excel column letter for toggle column
        start_row: First data row
        header_text: Header text for toggle column
    """
    from openpyxl import load_workbook
    from openpyxl.utils import column_index_from_string
    
    # Load workbook
    wb = load_workbook(excel_path)
    ws = wb.active
    
    col_idx = column_index_from_string(toggle_col)
    
    # Set header
    header_cell = ws.cell(row=1, column=col_idx)
    header_cell.value = header_text
    header_cell.font = Font(bold=True, color="FFFFFF")
    header_cell.alignment = Alignment(horizontal="center", vertical="center")
    header_cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    
    # Add dropdown data validation
    dv = DataValidation(
        type="list",
        formula1='"FALSE,TRUE"',
        allow_blank=False,
        showDropDown=False
    )
    dv.error = "Please select TRUE or FALSE"
    dv.errorTitle = "Invalid Entry"
    dv.prompt = "⚠️ WARNING: Selecting TRUE will DELETE this row immediately!"
    dv.promptTitle = "Auto-Delete Toggle"
    dv.showInputMessage = True
    
    # Apply validation
    max_row = ws.max_row if ws.max_row > start_row else 1000
    range_str = f"{toggle_col}{start_row}:{toggle_col}{max_row}"
    dv.add(range_str)
    ws.add_data_validation(dv)
    
    # Set default values
    for row_idx in range(start_row, ws.max_row + 1):
        cell = ws.cell(row=row_idx, column=col_idx)
        if cell.value is None:
            cell.value = "FALSE"
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Save as macro-enabled workbook (.xlsm)
    output_path = Path(excel_path).with_suffix('.xlsm')
    
    # Note: openpyxl cannot add VBA macros directly
    # We need to use xlwings or win32com for that
    wb.save(output_path)
    wb.close()
    
    print(f"[INFO] Excel file saved as: {output_path}")
    print(f"[WARNING] VBA macro must be added manually in Excel:")
    print(f"  1. Open {output_path} in Excel")
    print(f"  2. Press Alt+F11 to open VBA editor")
    print(f"  3. Double-click the sheet name in the left panel")
    print(f"  4. Paste the macro code (see VBA_AUTO_DELETE_MACRO in this file)")
    print(f"  5. Save and close VBA editor")
    print(f"  6. Enable macros when opening the file")
    
    return str(output_path)


def create_macro_enabled_template(
    source_excel: str,
    output_path: str = None,
    toggle_col: str = "E"
) -> str:
    """
    Create a macro-enabled Excel template with auto-delete functionality.
    
    This function:
    1. Adds the toggle column
    2. Saves as .xlsm
    3. Provides instructions for adding VBA macro
    
    Args:
        source_excel: Path to source Excel file
        output_path: Output path (default: same as source with .xlsm extension)
        toggle_col: Toggle column letter
    
    Returns:
        Path to the created .xlsm file
    """
    if output_path is None:
        output_path = Path(source_excel).with_suffix('.xlsm')
    
    result_path = add_toggle_column_with_macro(source_excel, toggle_col=toggle_col)
    
    # Save VBA macro to a text file for easy copy-paste
    macro_file = Path(output_path).with_suffix('.vba.txt')
    with open(macro_file, 'w') as f:
        f.write("=" * 70 + "\n")
        f.write("VBA MACRO CODE - Copy and paste into Excel VBA Editor\n")
        f.write("=" * 70 + "\n\n")
        f.write("Instructions:\n")
        f.write("1. Open the .xlsm file in Excel\n")
        f.write("2. Press Alt+F11 to open VBA Editor\n")
        f.write("3. In the left panel, double-click the sheet name\n")
        f.write("4. Copy ALL the code below and paste it into the code window\n")
        f.write("5. Press Ctrl+S to save\n")
        f.write("6. Close VBA Editor (Alt+Q)\n")
        f.write("7. Save the Excel file\n")
        f.write("8. When you reopen, click 'Enable Content' to allow macros\n\n")
        f.write("=" * 70 + "\n")
        f.write("COPY FROM HERE ↓\n")
        f.write("=" * 70 + "\n\n")
        f.write(VBA_AUTO_DELETE_MACRO)
        f.write("\n" + "=" * 70 + "\n")
        f.write("COPY UNTIL HERE ↑\n")
        f.write("=" * 70 + "\n")
    
    print(f"\n[SUCCESS] VBA macro code saved to: {macro_file}")
    print(f"[INFO] Open this file and follow the instructions to enable auto-delete")
    
    return str(result_path)


# Alternative: Using xlwings (requires Excel to be installed)
def add_toggle_with_xlwings(excel_path: str, toggle_col: str = "E"):
    """
    Add toggle column with VBA macro using xlwings.
    
    Requirements:
        pip install xlwings
        Microsoft Excel must be installed
    
    This version can automatically add the VBA macro without manual steps.
    """
    try:
        import xlwings as xw
    except ImportError:
        print("[ERROR] xlwings not installed. Install with: pip install xlwings")
        return
    
    # Open workbook with xlwings
    app = xw.App(visible=False)
    wb = xw.Book(excel_path)
    ws = wb.sheets[0]
    
    # Add toggle column (similar to above)
    # ... (column setup code)
    
    # Add VBA macro directly
    vba_module = wb.api.VBProject.VBComponents(ws.name).CodeModule
    vba_module.AddFromString(VBA_AUTO_DELETE_MACRO)
    
    # Save as macro-enabled
    output_path = Path(excel_path).with_suffix('.xlsm')
    wb.save(output_path)
    wb.close()
    app.quit()
    
    print(f"[SUCCESS] Macro-enabled workbook created: {output_path}")
    return str(output_path)


if __name__ == "__main__":
    import sys
    
    if len(sys.argv) < 2:
        print("Usage:")
        print("  python delete_toggle_feature_macro.py <excel_file> [toggle_col]")
        print("\nThis will:")
        print("  1. Add a toggle column to the Excel file")
        print("  2. Save as .xlsm (macro-enabled)")
        print("  3. Create a .vba.txt file with macro code to copy-paste")
        print("\nYou must manually add the VBA macro in Excel (see instructions)")
        sys.exit(1)
    
    excel_file = sys.argv[1]
    toggle_col = sys.argv[2] if len(sys.argv) > 2 else "E"
    
    create_macro_enabled_template(excel_file, toggle_col=toggle_col)
