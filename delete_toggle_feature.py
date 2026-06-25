# delete_toggle_feature.py
"""
Delete Toggle Feature Implementation

Provides functionality to:
1. Add a toggle/checkbox column to Excel output
2. Read toggle status from Excel files
3. Delete rows marked for deletion
"""

import openpyxl
from pathlib import Path
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.datavalidation import DataValidation


def add_toggle_column_to_workbook(
    ws,
    start_row: int = 2,
    toggle_col: str = "E",
    header_text: str = "Delete Toggle"
) -> None:
    """
    Add a toggle column with data validation to an existing worksheet.
    
    Args:
        ws: openpyxl worksheet object
        start_row: First data row (default 2, assuming row 1 is header)
        toggle_col: Excel column letter for the toggle column (default "E")
        header_text: Header text for the toggle column
    """
    from openpyxl.utils import column_index_from_string
    
    col_idx = column_index_from_string(toggle_col)
    
    # Set header
    header_cell = ws.cell(row=1, column=col_idx)
    header_cell.value = header_text
    header_cell.font = Font(bold=True)
    header_cell.alignment = Alignment(horizontal="center", vertical="center")
    header_cell.fill = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")
    
    # Add dropdown data validation for checkbox symbols
    dv = DataValidation(
        type="list",
        formula1='"☐,☑"',  # Unchecked and checked box symbols
        allow_blank=False,
        showDropDown=False
    )
    dv.error = "Please select ☐ (keep) or ☑ (delete)"
    dv.errorTitle = "Invalid Entry"
    dv.prompt = "Click to select: ☐ = Keep row, ☑ = Delete row"
    dv.promptTitle = "Delete Toggle"
    dv.showInputMessage = True
    
    # Apply validation to all data rows (up to row 1000 for safety)
    max_row = ws.max_row if ws.max_row > start_row else 1000
    range_str = f"{toggle_col}{start_row}:{toggle_col}{max_row}"
    dv.add(range_str)
    ws.add_data_validation(dv)
    
    # Set default value to unchecked box for existing rows
    for row_idx in range(start_row, ws.max_row + 1):
        cell = ws.cell(row=row_idx, column=col_idx)
        if cell.value is None:
            cell.value = "☐"  # Unchecked box symbol
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.font = Font(size=14)  # Larger font for better visibility


def read_toggle_status(
    excel_path: str,
    toggle_col: str = "E",
    start_row: int = 2,
    sheet_name: str | None = None
) -> dict[int, bool]:
    """
    Read toggle status from an Excel file.
    
    Args:
        excel_path: Path to the Excel file
        toggle_col: Excel column letter for the toggle column
        start_row: First data row (default 2)
        sheet_name: Target sheet name (None = first sheet)
    
    Returns:
        Dictionary mapping row numbers to toggle status (True = delete)
    """
    from openpyxl.utils import column_index_from_string
    
    wb = openpyxl.load_workbook(excel_path)
    ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active
    
    col_idx = column_index_from_string(toggle_col)
    toggle_status = {}
    
    for row_idx in range(start_row, ws.max_row + 1):
        cell_value = ws.cell(row=row_idx, column=col_idx).value
        # Consider checked box (☑), TRUE, True, true, 1, "1" as True
        is_delete = str(cell_value) in ("☑", "TRUE", "True", "true", "1") or str(cell_value).upper() == "TRUE"
        toggle_status[row_idx] = is_delete
    
    wb.close()
    return toggle_status


def delete_toggled_rows(
    excel_path: str,
    output_path: str | None = None,
    toggle_col: str = "E",
    start_row: int = 2,
    sheet_name: str | None = None,
    remove_toggle_column: bool = True
) -> tuple[int, str]:
    """
    Delete rows marked for deletion and save to a new file.
    
    Args:
        excel_path: Path to the source Excel file
        output_path: Path for the output file (None = overwrite source)
        toggle_col: Excel column letter for the toggle column
        start_row: First data row (default 2)
        sheet_name: Target sheet name (None = first sheet)
        remove_toggle_column: Whether to remove the toggle column after deletion
    
    Returns:
        Tuple of (deleted_count, output_file_path)
    """
    from openpyxl.utils import column_index_from_string
    
    if output_path is None:
        # Create backup and overwrite original
        backup_path = Path(excel_path).with_suffix('.backup.xlsx')
        import shutil
        shutil.copy2(excel_path, backup_path)
        output_path = excel_path
        print(f"[INFO] Backup created: {backup_path}")
    
    wb = openpyxl.load_workbook(excel_path)
    ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active
    
    col_idx = column_index_from_string(toggle_col)
    
    # Collect rows to delete (in reverse order to avoid index shifting)
    rows_to_delete = []
    for row_idx in range(start_row, ws.max_row + 1):
        cell_value = ws.cell(row=row_idx, column=col_idx).value
        # Check for checked box symbol or TRUE
        is_delete = str(cell_value) in ("☑", "TRUE", "True", "true", "1") or str(cell_value).upper() == "TRUE"
        if is_delete:
            rows_to_delete.append(row_idx)
    
    # Delete rows in reverse order
    deleted_count = 0
    for row_idx in reversed(rows_to_delete):
        ws.delete_rows(row_idx, 1)
        deleted_count += 1
    
    # Remove toggle column if requested
    if remove_toggle_column:
        ws.delete_cols(col_idx, 1)
        print(f"[INFO] Toggle column '{toggle_col}' removed")
    
    wb.save(output_path)
    wb.close()
    
    return deleted_count, output_path


def create_delete_toggle_summary(
    excel_path: str,
    toggle_col: str = "E",
    start_row: int = 2,
    sheet_name: str | None = None
) -> dict[str, int]:
    """
    Generate a summary of toggle status without modifying the file.
    
    Args:
        excel_path: Path to the Excel file
        toggle_col: Excel column letter for the toggle column
        start_row: First data row
        sheet_name: Target sheet name
    
    Returns:
        Dictionary with counts: {"total": n, "marked_for_deletion": m, "kept": k}
    """
    toggle_status = read_toggle_status(excel_path, toggle_col, start_row, sheet_name)
    
    total = len(toggle_status)
    marked = sum(1 for status in toggle_status.values() if status)
    kept = total - marked
    
    return {
        "total": total,
        "marked_for_deletion": marked,
        "kept": kept
    }


# CLI utility functions for standalone usage
def main_add_toggle_column(excel_path: str, toggle_col: str = "E") -> None:
    """Add toggle column to an existing Excel file."""
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active
    add_toggle_column_to_workbook(ws, toggle_col=toggle_col)
    wb.save(excel_path)
    print(f"[SUCCESS] Toggle column '{toggle_col}' added to {excel_path}")


def main_delete_toggled(
    excel_path: str,
    output_path: str | None = None,
    toggle_col: str = "E"
) -> None:
    """Delete toggled rows from Excel file."""
    count, out_path = delete_toggled_rows(
        excel_path,
        output_path,
        toggle_col=toggle_col
    )
    print(f"[SUCCESS] Deleted {count} row(s)")
    print(f"[SUCCESS] Output saved to: {out_path}")


def main_show_summary(excel_path: str, toggle_col: str = "E") -> None:
    """Show toggle status summary."""
    summary = create_delete_toggle_summary(excel_path, toggle_col=toggle_col)
    print(f"[INFO] Toggle Summary for: {excel_path}")
    print(f"  Total rows: {summary['total']}")
    print(f"  Marked for deletion: {summary['marked_for_deletion']}")
    print(f"  Will be kept: {summary['kept']}")


if __name__ == "__main__":
    import sys
    
    if len(sys.argv) < 2:
        print("Usage:")
        print("  Add toggle column:    python delete_toggle_feature.py add <excel_file> [toggle_col]")
        print("  Delete toggled rows:  python delete_toggle_feature.py delete <excel_file> [output_file] [toggle_col]")
        print("  Show summary:         python delete_toggle_feature.py summary <excel_file> [toggle_col]")
        sys.exit(1)
    
    command = sys.argv[1].lower()
    
    if command == "add":
        excel_file = sys.argv[2]
        toggle_col = sys.argv[3] if len(sys.argv) > 3 else "E"
        main_add_toggle_column(excel_file, toggle_col)
    
    elif command == "delete":
        excel_file = sys.argv[2]
        output_file = sys.argv[3] if len(sys.argv) > 3 else None
        toggle_col = sys.argv[4] if len(sys.argv) > 4 else "E"
        main_delete_toggled(excel_file, output_file, toggle_col)
    
    elif command == "summary":
        excel_file = sys.argv[2]
        toggle_col = sys.argv[3] if len(sys.argv) > 3 else "E"
        main_show_summary(excel_file, toggle_col)
    
    else:
        print(f"Unknown command: {command}")
        sys.exit(1)
