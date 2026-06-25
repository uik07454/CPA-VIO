# gui.py
"""PyQt6 GUI entry point for the compare_crs_docx pipeline."""

import sys
from pathlib import Path

from PyQt6.QtCore import Qt, QThread, pyqtSignal
from PyQt6.QtGui import QColor, QIcon, QPalette
from PyQt6.QtWidgets import (
    QApplication,
    QCheckBox,
    QColorDialog,
    QComboBox,
    QDialog,
    QDialogButtonBox,
    QFileDialog,
    QFormLayout,
    QGridLayout,
    QGroupBox,
    QHBoxLayout,
    QHeaderView,
    QLabel,
    QLineEdit,
    QMainWindow,
    QMessageBox,
    QPushButton,
    QRadioButton,
    QScrollArea,
    QSlider,
    QSplitter,
    QStatusBar,
    QTableWidget,
    QTableWidgetItem,
    QTreeWidget,
    QTreeWidgetItem,
    QVBoxLayout,
    QWidget,
)

from PyQt6.QtMultimedia import QMediaPlayer, QAudioOutput
from PyQt6.QtMultimediaWidgets import QVideoWidget
from PyQt6.QtCore import QUrl

from app_config import AppConfig
from constants import (
    APP_ICON_PATH,
    BASE_DIR, CPA_TEMPLATE_DIR,
    CHANGE_TYPE_CATEGORIES,
    IMAGE_PX_MAX, IMAGE_PX_MIN,
    IMAGE_SIZE_MODE_FIXED, IMAGE_SIZE_MODE_ORIGINAL,
    OUTPUT_MODE_SCRATCH, OUTPUT_MODE_TEMPLATE,
    SPLASH_VIDEO_PATH,
)


# ---------------------------------------------------------------------------
# Splash screen
# ---------------------------------------------------------------------------

class _SplashScreen(QWidget):
    """
    Frameless fullscreen splash that plays assets/splash.mp4.

    Closes itself when:
      - Playback ends naturally, OR
      - The user clicks anywhere on the screen (skip).
    Emits no signals; the caller connects to ``finished`` to know when to show
    the main window.
    """

    finished = pyqtSignal()

    def __init__(self) -> None:
        super().__init__(
            None,
            Qt.WindowType.FramelessWindowHint | Qt.WindowType.WindowStaysOnTopHint,
        )
        self.setAttribute(Qt.WidgetAttribute.WA_DeleteOnClose)
        self.setStyleSheet("background: black;")

        layout = QVBoxLayout(self)
        layout.setContentsMargins(0, 0, 0, 0)

        self._video = QVideoWidget()
        layout.addWidget(self._video)

        self._player = QMediaPlayer()
        self._audio  = QAudioOutput()
        self._player.setAudioOutput(self._audio)
        self._player.setVideoOutput(self._video)
        self._player.playbackStateChanged.connect(self._on_state_changed)

        video_path = str(SPLASH_VIDEO_PATH.resolve())
        self._player.setSource(QUrl.fromLocalFile(video_path))

        if APP_ICON_PATH.exists():
            self.setWindowIcon(QIcon(str(APP_ICON_PATH)))

    def show_and_play(self) -> None:
        """Show fullscreen and start playback."""
        self.showFullScreen()
        self._player.play()

    def mousePressEvent(self, event) -> None:
        """Allow user to skip the splash by clicking."""
        self._finish()

    def _on_state_changed(self, state: QMediaPlayer.PlaybackState) -> None:
        if state == QMediaPlayer.PlaybackState.StoppedState:
            self._finish()

    def _finish(self) -> None:
        self._player.stop()
        self.close()
        self.finished.emit()


# ---------------------------------------------------------------------------
# Worker thread
# ---------------------------------------------------------------------------

class _PipelineWorker(QThread):
    """Runs the extraction pipeline off the main thread."""

    log     = pyqtSignal(str)
    finished = pyqtSignal(bool, str)   # success, message

    def __init__(self, docx_path: str, config: AppConfig) -> None:
        super().__init__()
        self._docx  = docx_path
        self._cfg   = config

    def run(self) -> None:
        try:
            from parser import (
                parse_any_xml, build_outline_level_map,
                build_image_map, parse_headers_footers, save_xml_to_output,
            )
            from extractor import extract_tracked_changes, extract_hf_changes
            from book_marker import inject_bookmarks
            from output import generate_excel_output, generate_excel_from_template
            from constants import CPA_TEMPLATE_DIR, OUTPUT_MODE_TEMPLATE, TO_PARSED_XML_PARTS

            # For developer, manually toggle this for saving XML for debugging
            _SAVE_XML = False 

            self.log.emit(f"[INFO] Reading: {self._docx}")
            parse_results  = parse_any_xml(self._docx, TO_PARSED_XML_PARTS, save_xml=_SAVE_XML)
            doc_root       = parse_results["document"]
            styles_root    = parse_results["styles"]
            footnotes_root = parse_results["footnotes"]
            endnotes_root  = parse_results["endnotes"]

            self.log.emit("[INFO] Building outline level map ...")
            outline_map = build_outline_level_map(styles_root)

            self.log.emit("[INFO] Building image map ...")
            image_map = build_image_map(self._docx)

            note_roots = {"footnote": footnotes_root, "endnote": endnotes_root}

            self.log.emit("[INFO] Extracting tracked changes ...")
            records = extract_tracked_changes(
                doc_root, outline_map, include_sdt=self._cfg.include_sdt,
                image_map=image_map, note_roots=note_roots,
                default_heading=self._cfg.default_heading,
            )
            self.log.emit(f"[INFO] Found {len(records)} document body change(s).")

            hf_records = []
            if self._cfg.include_hf_changes:
                hf_roots   = parse_headers_footers(self._docx, save_xml=_SAVE_XML)
                hf_records = extract_hf_changes(hf_roots)
                self.log.emit(f"[INFO] Found {len(hf_records)} header/footer change(s).")

            all_records = self._cfg.filter_records(hf_records + records)
            self.log.emit(f"[INFO] {len(all_records)} record(s) after change-type filter.")

            output_dir = BASE_DIR / "output"
            output_dir.mkdir(exist_ok=True)
            annotated_path = str(output_dir / (Path(self._docx).stem + "_annotated.docx"))
            annotated_root = inject_bookmarks(self._docx, all_records, annotated_path)

            if _SAVE_XML:
                save_xml_to_output(annotated_root, annotated_path, suffix="document")
                self.log.emit(f"[INFO] Annotated document XML saved to output/")

            excel_out = str(output_dir / (Path(self._docx).stem + "_changes.xlsx"))

            if self._cfg.output_mode == OUTPUT_MODE_TEMPLATE and self._cfg.selected_template:
                template_path  = str(CPA_TEMPLATE_DIR / self._cfg.selected_template)
                column_mapping = self._cfg.get_column_mapping()
                self.log.emit(f"[INFO] Using template: {self._cfg.selected_template}")
                sheet_name = self._cfg.get_selected_sheet()
                self.log.emit(f"[INFO] Writing to sheet: '{sheet_name}' (fallback to first sheet if not found)")
                generate_excel_from_template(
                    all_records, annotated_path, excel_out,
                    template_path=template_path,
                    column_mapping=column_mapping,
                    config=self._cfg,
                    sheet_name=sheet_name,
                )
            else:
                generate_excel_output(all_records, annotated_path, excel_out, config=self._cfg)

            # Add delete toggle column if enabled
            if self._cfg.enable_delete_toggle:
                from delete_toggle_feature import add_toggle_column_to_workbook
                import openpyxl
                
                self.log.emit(f"[INFO] Adding delete toggle column '{self._cfg.delete_toggle_column}' ...")
                wb = openpyxl.load_workbook(excel_out)
                ws = wb.active
                add_toggle_column_to_workbook(
                    ws,
                    start_row=2,
                    toggle_col=self._cfg.delete_toggle_column,
                    header_text="Delete Toggle"
                )
                wb.save(excel_out)
                wb.close()
                self.log.emit(f"[INFO] Delete toggle column added successfully")

            self.finished.emit(True, f"Done! Excel saved to: {excel_out}")
        except Exception as exc:
            self.finished.emit(False, f"[ERROR] {exc}")


# ---------------------------------------------------------------------------
# Color swatch button
# ---------------------------------------------------------------------------

class _ColorButton(QPushButton):
    """Square button showing the current colour; opens a colour picker on click."""

    def __init__(self, hex_color: str, parent: QWidget | None = None) -> None:
        super().__init__(parent)
        self.setFixedSize(28, 28)
        self.set_color(hex_color)
        self.clicked.connect(self._pick)

    def set_color(self, hex_color: str) -> None:
        self._hex = hex_color.upper()
        
        pal = self.palette()
        color = QColor(f"#{self._hex}")
        pal.setColor(QPalette.ColorRole.Button, color)

        text_color = QColor("#000000") if color.lightness() > 128 else QColor("#FFFFFF")
        pal.setColor(QPalette.ColorRole.ButtonText, text_color)

        self.setPalette(pal)
        self.setAutoFillBackground(True)
        self.update()

    def hex_color(self) -> str:
        return self._hex

    def _pick(self) -> None:
        initial = QColor(f"#{self._hex}")
        chosen  = QColorDialog.getColor(initial, self, "Select Colour")
        if chosen.isValid():
            self.set_color(chosen.name().lstrip("#"))
            # Notify linked QLineEdit if stored as sibling attribute
            if hasattr(self, "_linked_edit"):
                self._linked_edit.setText(self._hex)


# ---------------------------------------------------------------------------
# Settings panel (left side)
# ---------------------------------------------------------------------------

class _SettingsPanel(QWidget):
    """Left panel: Text, Image, and Diff Color settings groups."""

    def __init__(self, config: AppConfig, parent: QWidget | None = None) -> None:
        super().__init__(parent)
        self._cfg = config
        self._build_ui()

    def _build_ui(self) -> None:
        layout = QVBoxLayout(self)
        layout.setContentsMargins(8, 8, 8, 8)
        layout.addWidget(self._make_text_group())
        layout.addWidget(self._make_image_group())
        layout.addWidget(self._make_color_group())
        self._output_panel = _OutputSettingsPanel(self._cfg)
        layout.addWidget(self._output_panel)
        layout.addWidget(self._make_toggle_group())
        layout.addStretch()

    def _make_text_group(self) -> QGroupBox:
        grp    = QGroupBox("Text Settings")
        layout = QVBoxLayout(grp)

        self._empty_field_edit   = QLineEdit()
        self._default_heading_edit = QLineEdit()

        for lbl, widget, tip in (
            ("Empty Field Placeholder:", self._empty_field_edit, "Text to be display for empty field content"),
            ("Default Heading:", self._default_heading_edit, "Heading text to be display if the change record doesn't have heading"),
        ):
            row = QHBoxLayout()
            qlbl = QLabel(lbl)
            row.addWidget(qlbl)
            qlbl.setToolTip(tip)
            qlbl.setFixedWidth(140)
            row.addWidget(widget)
            layout.addLayout(row)

        return grp

    def _make_image_group(self) -> QGroupBox:
        grp    = QGroupBox("Image Settings")
        layout = QVBoxLayout(grp)

        # ---- Mode radio buttons ----
        mode_row = QHBoxLayout()
        self._mode_fixed_radio    = QRadioButton("Fixed Size")
        self._mode_original_radio = QRadioButton("Image Original Size")
        self._mode_fixed_radio.setToolTip("All images use the configured width/height")
        self._mode_original_radio.setToolTip(
            "Each image uses its own size from the document.\n"
            "Fallback to configured size if original size is unavailable."
        )
        self._mode_original_radio.setChecked(True)

        mode_row.addWidget(self._mode_original_radio)
        mode_row.addWidget(self._mode_fixed_radio)
        mode_row.addStretch()
        layout.addLayout(mode_row)

        # ---- Width and height sliders ----
        self._img_w_slider, self._img_w_label = self._make_px_slider()
        self._img_w_slider.valueChanged.connect(
            lambda v: self._img_w_label.setText(f"{v} px")
        )

        self._img_h_slider, self._img_h_label = self._make_px_slider()
        self._img_h_slider.valueChanged.connect(self._update_row_height_display)

        # ---- Row height display (read-only label) ----
        self._row_h_label = QLabel()
        self._row_h_label.setToolTip("Auto-calculated row height based on image height")
        self._row_h_label.setStyleSheet("color: gray;")

        # Use grid layout for better alignment
        grid = QGridLayout()

        # Add sliders and labels to the grid
        for i, (lbl_text, slider, val_label) in enumerate(
            [("Default Image Width :", self._img_w_slider, self._img_w_label),
             ("Default Image Height:", self._img_h_slider, self._img_h_label)],
        ):
            self._image_text_label = QLabel(lbl_text)
            self._image_text_label.setToolTip("Default dimensions for images in Excel output.\n")
            grid.addWidget(self._image_text_label, i, 0)
            grid.addWidget(slider,                 i, 1)
            grid.addWidget(val_label,              i, 2)

        # Add the auto-calculated row height display below the sliders
        grid.addWidget(QLabel("Image Row Height:"), 2, 0)
        grid.addWidget(self._row_h_label, 2, 2)

        layout.addLayout(grid)

        return grp

    def _make_px_slider(self) -> tuple[QSlider, QLabel]:
        """Return a horizontal pixel slider paired with a value display label."""
        slider = QSlider(Qt.Orientation.Horizontal)
        slider.setRange(IMAGE_PX_MIN, IMAGE_PX_MAX)
        slider.setSingleStep(10)
        slider.setPageStep(50)
        label = QLabel(f"{slider.value()} px")
        label.setFixedWidth(35)
        return slider, label

    def _update_row_height_display(self, height_px: int | None = None) -> None:
        if height_px is None:
            height_px = self._img_h_slider.value()
        self._img_h_label.setText(f"{height_px} px")
        row_h_pt = int(height_px * 0.75) + 20
        self._row_h_label.setText(f"{row_h_pt} pt")

    def _make_color_group(self) -> QGroupBox:
        grp    = QGroupBox("Diff Colors (apply to modified content-changes only)")
        layout = QVBoxLayout(grp)

        del_row, self._del_btn, self._del_edit = self._make_color_row("Delete Color:", self._cfg.diff_del_color)
        ins_row, self._ins_btn, self._ins_edit = self._make_color_row("Insert Color:", self._cfg.diff_ins_color)
        eq_row,  self._eq_btn,  self._eq_edit  = self._make_color_row("Equal Color:",  self._cfg.diff_eq_color)

        layout.addLayout(del_row)
        layout.addLayout(ins_row)
        layout.addLayout(eq_row)
        return grp

    def _make_color_row(self, label: str, hex_val: str) -> tuple[QHBoxLayout, _ColorButton, QLineEdit]:
        row  = QHBoxLayout()
        btn  = _ColorButton(hex_val)
        edit = QLineEdit(hex_val)
        edit.setFixedWidth(70)
        edit.setReadOnly(True)

        btn._linked_edit = edit

        qlbl = QLabel(label)
        row.addWidget(qlbl)
        qlbl.setFixedWidth(75)
        row.addWidget(btn)
        row.addWidget(edit)
        row.addStretch()
        return row, btn, edit

    def _make_toggle_group(self) -> QGroupBox:
        grp = QGroupBox("Delete Toggle Feature")
        layout = QVBoxLayout(grp)

        self._enable_toggle_checkbox = QCheckBox("Enable Delete Toggle Column")
        self._enable_toggle_checkbox.setToolTip(
            "Add a toggle column to Excel output for marking rows to delete.\n"
            "Users can set toggle to TRUE/FALSE, then use 'Delete Toggled Rows' to remove marked rows."
        )
        layout.addWidget(self._enable_toggle_checkbox)

        col_row = QHBoxLayout()
        col_row.addWidget(QLabel("Toggle Column:"))
        self._toggle_col_edit = QLineEdit("E")
        self._toggle_col_edit.setMaxLength(3)
        self._toggle_col_edit.setFixedWidth(50)
        self._toggle_col_edit.setToolTip("Excel column letter for the delete toggle column (e.g., E, F, Z)")
        col_row.addWidget(self._toggle_col_edit)
        col_row.addStretch()
        layout.addLayout(col_row)

        self._delete_toggled_btn = QPushButton("Delete Toggled Rows from Excel...")
        self._delete_toggled_btn.setToolTip("Open an Excel file and delete rows marked with TRUE in the toggle column")
        self._delete_toggled_btn.clicked.connect(self._on_delete_toggled_rows)
        layout.addWidget(self._delete_toggled_btn)

        return grp

    def _on_delete_toggled_rows(self) -> None:
        """Handle the 'Delete Toggled Rows' button click."""
        from delete_toggle_feature import delete_toggled_rows, create_delete_toggle_summary

        excel_path, _ = QFileDialog.getOpenFileName(
            self, "Select Excel File with Toggle Column", "", "Excel Files (*.xlsx)"
        )
        if not excel_path:
            return

        try:
            # Show summary first
            toggle_col = self._toggle_col_edit.text().strip().upper() or "E"
            summary = create_delete_toggle_summary(excel_path, toggle_col=toggle_col)
            
            if summary["marked_for_deletion"] == 0:
                QMessageBox.information(
                    self, "No Rows to Delete",
                    f"No rows are marked for deletion (toggle = TRUE).\n\n"
                    f"Total rows: {summary['total']}"
                )
                return

            # Confirm deletion
            reply = QMessageBox.question(
                self, "Confirm Deletion",
                f"Delete {summary['marked_for_deletion']} row(s) marked with TRUE?\n\n"
                f"Total rows: {summary['total']}\n"
                f"Marked for deletion: {summary['marked_for_deletion']}\n"
                f"Will be kept: {summary['kept']}\n\n"
                f"A backup file will be created automatically.",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
            )

            if reply != QMessageBox.StandardButton.Yes:
                return

            # Perform deletion
            deleted_count, output_path = delete_toggled_rows(
                excel_path,
                toggle_col=toggle_col,
                remove_toggle_column=True
            )

            QMessageBox.information(
                self, "Deletion Complete",
                f"Successfully deleted {deleted_count} row(s).\n\n"
                f"Output saved to: {output_path}\n"
                f"Backup created with .backup.xlsx extension"
            )

        except Exception as e:
            QMessageBox.critical(
                self, "Error",
                f"Failed to delete toggled rows:\n{str(e)}"
            )

    def apply_to_config(self) -> None:
        self._cfg.empty_field         = self._empty_field_edit.text()
        self._cfg.default_heading     = self._default_heading_edit.text()
        self._cfg.image_width_px      = self._img_w_slider.value()
        self._cfg.image_height_px     = self._img_h_slider.value()
        self._cfg.image_size_mode     = (
            IMAGE_SIZE_MODE_FIXED if self._mode_fixed_radio.isChecked()
            else IMAGE_SIZE_MODE_ORIGINAL
        )
        self._cfg.diff_del_color      = self._del_edit.text().upper()
        self._cfg.diff_ins_color      = self._ins_edit.text().upper()
        self._cfg.diff_eq_color       = self._eq_edit.text().upper()
        self._cfg.enable_delete_toggle = self._enable_toggle_checkbox.isChecked()
        self._cfg.delete_toggle_column = self._toggle_col_edit.text().strip().upper() or "E"
        self._output_panel.apply_to_config()

    def is_image_fixed_mode(self) -> bool:
        return self._cfg.image_size_mode == IMAGE_SIZE_MODE_FIXED

    def load_from_config(self) -> None:
        self._empty_field_edit.setText(self._cfg.empty_field)
        self._default_heading_edit.setText(self._cfg.default_heading)
        self._img_w_slider.setValue(self._cfg.image_width_px)
        self._img_h_slider.setValue(self._cfg.image_height_px)
        self._mode_fixed_radio.setChecked(self.is_image_fixed_mode())
        self._mode_original_radio.setChecked(not self.is_image_fixed_mode())
        self._update_row_height_display()
        self._del_edit.setText(self._cfg.diff_del_color)
        self._ins_edit.setText(self._cfg.diff_ins_color)
        self._eq_edit.setText(self._cfg.diff_eq_color)
        self._del_btn.set_color(self._cfg.diff_del_color)
        self._ins_btn.set_color(self._cfg.diff_ins_color)
        self._eq_btn.set_color(self._cfg.diff_eq_color)
        self._enable_toggle_checkbox.setChecked(self._cfg.enable_delete_toggle)
        self._toggle_col_edit.setText(self._cfg.delete_toggle_column)
        self._output_panel.load_from_config()

    def validation_errors(self) -> list[str]:
        errors: list[str] = []
        if not self._empty_field_edit.text():
            errors.append("Empty Field Placeholder must not be empty.")
        if not self._default_heading_edit.text():
            errors.append("Default Heading must not be empty.")
        errors.extend(self._output_panel.validation_errors())
        return errors


# ---------------------------------------------------------------------------
# Change type tree (right side)
# ---------------------------------------------------------------------------

class _ChangeTypeTree(QWidget):
    """Right panel: hierarchical checkbox tree for change type selection."""

    def __init__(self, config: AppConfig, parent: QWidget | None = None) -> None:
        super().__init__(parent)
        self._cfg = config
        self._category_items: dict[str, QTreeWidgetItem] = {}
        self._build_ui()

    def _build_ui(self) -> None:
        layout = QVBoxLayout(self)
        layout.setContentsMargins(8, 8, 8, 8)

        # Include header/footer changes toggle
        self._hf_checkbox = QCheckBox("Include Header/Footer Changes")
        self._hf_checkbox.setToolTip(
            "Output each page Header/Footer differences individually even if they share the same difference\n\n"
            "Recommended: Manually check the Header/Footer change instead of using this option"
        )
        self._hf_checkbox.setChecked(True)
        layout.addWidget(self._hf_checkbox)

        # Include SDT (content control) changes toggle
        self._sdt_checkbox = QCheckBox("Include SDT (Content Control) Changes")
        self._sdt_checkbox.setToolTip(
            "SDT (Structured Document Tags) are Word content controls such as\n"
            "drop-down lists, rich text boxes, and Table of Contents (<w:sdt> elements).\n"
            "When enabled, tracked changes inside these controls are also extracted.\n\n"
            "TLDR: SDT = auto-calculated content by Word (mostly can be ignored, not a cp)"
        )
        self._sdt_checkbox.setChecked(True)
        layout.addWidget(self._sdt_checkbox)

        # Select All / Deselect All buttons
        btn_row = QHBoxLayout()
        sel_all_btn = QPushButton("Select All")
        desel_btn   = QPushButton("Deselect All")
        sel_all_btn.clicked.connect(self.select_all)
        desel_btn.clicked.connect(self.deselect_all)
        btn_row.addWidget(sel_all_btn)
        btn_row.addWidget(desel_btn)
        btn_row.addStretch()
        layout.addLayout(btn_row)

        self._tree = QTreeWidget()
        self._tree.setHeaderHidden(True)
        self._tree.setColumnCount(1)

        for category, change_types in CHANGE_TYPE_CATEGORIES.items():
            cat_item = QTreeWidgetItem(self._tree, [category])
            cat_item.setFlags(cat_item.flags() | Qt.ItemFlag.ItemIsUserCheckable | Qt.ItemFlag.ItemIsAutoTristate)
            cat_item.setCheckState(0, Qt.CheckState.Checked)
            self._category_items[category] = cat_item

            for ct in change_types:
                child = QTreeWidgetItem(cat_item, [ct])
                child.setFlags(child.flags() | Qt.ItemFlag.ItemIsUserCheckable)
                child.setCheckState(0, Qt.CheckState.Checked)
                child.setData(0, Qt.ItemDataRole.UserRole, ct)

        self._tree.expandAll()
        self._tree.itemChanged.connect(self._on_item_changed)
        layout.addWidget(self._tree)

    def _on_item_changed(self, item: QTreeWidgetItem, column: int) -> None:
        # Propagate parent check state to all children
        if item.childCount() > 0:
            state = item.checkState(0)
            if state == Qt.CheckState.PartiallyChecked:
                return
            self._tree.blockSignals(True)
            for i in range(item.childCount()):
                item.child(i).setCheckState(0, state)
            self._tree.blockSignals(False)
        else:
            parent = item.parent()
            if parent:
                self._update_parent_state(parent)

    def _update_parent_state(self, parent: QTreeWidgetItem) -> None:
        checked = sum(
            1 for i in range(parent.childCount())
            if parent.child(i).checkState(0) == Qt.CheckState.Checked
        )
        total = parent.childCount()
        self._tree.blockSignals(True)
        if checked == 0:
            parent.setCheckState(0, Qt.CheckState.Unchecked)
        elif checked == total:
            parent.setCheckState(0, Qt.CheckState.Checked)
        else:
            parent.setCheckState(0, Qt.CheckState.PartiallyChecked)
        self._tree.blockSignals(False)

    def select_all(self) -> None:
        self._tree.blockSignals(True)
        for i in range(self._tree.topLevelItemCount()):
            cat = self._tree.topLevelItem(i)
            cat.setCheckState(0, Qt.CheckState.Checked)
            for j in range(cat.childCount()):
                cat.child(j).setCheckState(0, Qt.CheckState.Checked)
        self._tree.blockSignals(False)

    def deselect_all(self) -> None:
        self._tree.blockSignals(True)
        for i in range(self._tree.topLevelItemCount()):
            cat = self._tree.topLevelItem(i)
            cat.setCheckState(0, Qt.CheckState.Unchecked)
            for j in range(cat.childCount()):
                cat.child(j).setCheckState(0, Qt.CheckState.Unchecked)
        self._tree.blockSignals(False)

    def apply_to_config(self) -> None:
        self._cfg.include_hf_changes = self._hf_checkbox.isChecked()
        self._cfg.include_sdt        = self._sdt_checkbox.isChecked()
        selected: set[str] = set()
        for i in range(self._tree.topLevelItemCount()):
            cat = self._tree.topLevelItem(i)
            for j in range(cat.childCount()):
                child = cat.child(j)
                if child.checkState(0) == Qt.CheckState.Checked:
                    selected.add(child.data(0, Qt.ItemDataRole.UserRole))
        self._cfg.selected_change_types = selected

    def load_from_config(self) -> None:
        self._hf_checkbox.setChecked(self._cfg.include_hf_changes)
        self._sdt_checkbox.setChecked(self._cfg.include_sdt)
        self._tree.blockSignals(True)
        for i in range(self._tree.topLevelItemCount()):
            cat = self._tree.topLevelItem(i)
            for j in range(cat.childCount()):
                child = cat.child(j)
                ct    = child.data(0, Qt.ItemDataRole.UserRole)
                state = Qt.CheckState.Checked if ct in self._cfg.selected_change_types else Qt.CheckState.Unchecked
                child.setCheckState(0, state)
            self._update_parent_state(cat)
        self._tree.blockSignals(False)


# ---------------------------------------------------------------------------
# Template profile manager dialog
# ---------------------------------------------------------------------------

# Human-readable labels for the 4 output fields (matches OUTPUT_FIELDS order).
_FIELD_LABELS = ("Heading", "Change Type", "Old Content", "New Content")


def _read_sheet_names(template_name: str) -> list[str]:
    """Return sheet names from the given template file, or [] on failure."""
    import openpyxl
    path = CPA_TEMPLATE_DIR / template_name
    if not path.exists():
        return []
    try:
        wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
        names = wb.sheetnames
        wb.close()
        return names
    except Exception:
        return []


class _TemplateProfileDialog(QDialog):
    """Helper dialog to add, edit, or delete per-sheet column-mapping profiles."""

    def __init__(self, config: AppConfig, parent: QWidget | None = None) -> None:
        super().__init__(parent)
        self._cfg = config
        self.setWindowTitle("Manage Template Profiles")
        self.resize(720, 360)
        if APP_ICON_PATH.exists():
            self.setWindowIcon(QIcon(str(APP_ICON_PATH)))
        self._build_ui()
        self._refresh_table()

    def _build_ui(self) -> None:
        layout = QVBoxLayout(self)

        # Table: Template | Sheet | Heading col | Change Type col | Old col | New col
        self._table = QTableWidget(0, 6)
        self._table.setHorizontalHeaderLabels(
            ["Template File", "Sheet"] + list(_FIELD_LABELS)
        )
        self._table.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeMode.Stretch)
        self._table.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeMode.ResizeToContents)
        for c in range(2, 6):
            self._table.horizontalHeader().setSectionResizeMode(c, QHeaderView.ResizeMode.ResizeToContents)
        self._table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self._table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        layout.addWidget(self._table)

        btn_row = QHBoxLayout()
        self._add_btn  = QPushButton("Add…")
        self._edit_btn = QPushButton("Edit…")
        self._del_btn  = QPushButton("Delete")
        self._add_btn.clicked.connect(self._on_add)
        self._edit_btn.clicked.connect(self._on_edit)
        self._del_btn.clicked.connect(self._on_delete)
        btn_row.addWidget(self._add_btn)
        btn_row.addWidget(self._edit_btn)
        btn_row.addWidget(self._del_btn)
        btn_row.addStretch()
        layout.addLayout(btn_row)

        buttons = QDialogButtonBox(QDialogButtonBox.StandardButton.Close)
        buttons.rejected.connect(self.accept)
        layout.addWidget(buttons)

    def _refresh_table(self) -> None:
        self._table.setRowCount(0)
        for tpl_name, sheet_map in sorted(self._cfg.template_profiles.items()):
            for sheet_name, cols in sorted(sheet_map.items()):
                row = self._table.rowCount()
                self._table.insertRow(row)
                self._table.setItem(row, 0, QTableWidgetItem(tpl_name))
                self._table.setItem(row, 1, QTableWidgetItem(sheet_name))
                for c, val in enumerate(cols[:4], start=2):
                    self._table.setItem(row, c, QTableWidgetItem(val))

    def _open_profile_editor(
        self, tpl_name: str = "", sheet_name: str = "", cols: list[str] | None = None
    ) -> None:
        """Open an editor dialog for adding or editing a (template, sheet) profile."""
        dlg = QDialog(self)
        dlg.setWindowTitle("Edit Profile" if tpl_name else "Add Profile")
        dlg.resize(420, 240)
        form = QFormLayout(dlg)

        templates = self._cfg.available_templates()
        name_combo = QComboBox()
        name_combo.addItems(templates)
        if tpl_name and tpl_name in templates:
            name_combo.setCurrentText(tpl_name)
        name_combo.setEditable(False)
        form.addRow("Template:", name_combo)

        sheet_combo = QComboBox()
        sheet_combo.setEditable(False)

        def _refresh_sheets(selected_tpl: str) -> None:
            current_sheet = sheet_combo.currentText()
            sheet_combo.blockSignals(True)
            sheet_combo.clear()
            sheet_combo.addItems(_read_sheet_names(selected_tpl))
            idx = sheet_combo.findText(current_sheet)
            if idx >= 0:
                sheet_combo.setCurrentIndex(idx)
            sheet_combo.blockSignals(False)

        name_combo.currentTextChanged.connect(_refresh_sheets)
        _refresh_sheets(name_combo.currentText())
        if sheet_name:
            idx = sheet_combo.findText(sheet_name)
            if idx >= 0:
                sheet_combo.setCurrentIndex(idx)
        form.addRow("Sheet:", sheet_combo)

        col_edits: list[QLineEdit] = []
        defaults = cols or ["A", "B", "C", "D"]
        for label, default in zip(_FIELD_LABELS, defaults):
            edit = QLineEdit(default)
            edit.setMaxLength(3)
            edit.setFixedWidth(50)
            form.addRow(f"{label} column:", edit)
            col_edits.append(edit)

        btns = QDialogButtonBox(
            QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel
        )
        btns.accepted.connect(dlg.accept)
        btns.rejected.connect(dlg.reject)
        form.addRow(btns)

        if dlg.exec() != QDialog.DialogCode.Accepted:
            return

        chosen_tpl   = name_combo.currentText()
        chosen_sheet = sheet_combo.currentText()
        chosen_cols  = [e.text().strip().upper() for e in col_edits]

        if not chosen_sheet:
            QMessageBox.warning(self, "No Sheet", "No sheet selected. The template may have no sheets.")
            return

        import openpyxl.utils as _oxl_utils
        invalid = []
        for c in chosen_cols:
            try:
                if _oxl_utils.column_index_from_string(c) > 16384:  # XFD is the Excel max
                    raise ValueError
            except ValueError:
                invalid.append(c)
        if invalid:
            QMessageBox.warning(self, "Invalid Columns", f"Invalid column letters: {', '.join(invalid)}")
            return

        self._cfg.template_profiles.setdefault(chosen_tpl, {})[chosen_sheet] = chosen_cols
        self._refresh_table()

    def _on_add(self) -> None:
        self._open_profile_editor()

    def _on_edit(self) -> None:
        row = self._table.currentRow()
        if row < 0:
            QMessageBox.information(self, "No Selection", "Select a profile to edit.")
            return
        tpl_name   = self._table.item(row, 0).text()
        sheet_name = self._table.item(row, 1).text()
        cols = self._cfg.template_profiles.get(tpl_name, {}).get(sheet_name, ["A", "B", "C", "D"])
        self._open_profile_editor(tpl_name, sheet_name, cols)

    def _on_delete(self) -> None:
        row = self._table.currentRow()
        if row < 0:
            QMessageBox.information(self, "No Selection", "Select a profile to delete.")
            return
        tpl_name   = self._table.item(row, 0).text()
        sheet_name = self._table.item(row, 1).text()
        sheet_map  = self._cfg.template_profiles.get(tpl_name, {})
        sheet_map.pop(sheet_name, None)
        if not sheet_map:
            self._cfg.template_profiles.pop(tpl_name, None)
        self._refresh_table()


# ---------------------------------------------------------------------------
# Output settings panel (template / scratch selection)
# ---------------------------------------------------------------------------

class _OutputSettingsPanel(QWidget):
    """Group box for selecting output mode, template, and column mapping."""

    def __init__(self, config: AppConfig, parent: QWidget | None = None) -> None:
        super().__init__(parent)
        self._cfg = config
        self._build_ui()

    def _build_ui(self) -> None:
        layout = QVBoxLayout(self)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.addWidget(self._make_output_group())

    def _make_output_group(self) -> QGroupBox:
        grp    = QGroupBox("Output Settings")
        layout = QVBoxLayout(grp)

        # ---- Mode radio buttons ----
        mode_row = QHBoxLayout()
        self._scratch_radio  = QRadioButton("Generate from Scratch")
        self._scratch_radio.setToolTip("Output change records to a new .xlsx file.")
        self._template_radio = QRadioButton("Generate from Template")
        self._template_radio.setToolTip("Output change records to a CPA_template.xlsx file.")
        self._scratch_radio.setChecked(True)
        self._scratch_radio.toggled.connect(self._on_mode_toggled)
        mode_row.addWidget(self._scratch_radio)
        mode_row.addWidget(self._template_radio)
        mode_row.addStretch()
        layout.addLayout(mode_row)

        # ---- Template selector ----
        tpl_row = QHBoxLayout()
        tpl_lbl = QLabel("Template:")
        tpl_row.addWidget(tpl_lbl)
        tpl_lbl.setFixedWidth(55)           # smooth operatorr
        self._template_combo = QComboBox()
        self._template_combo.currentTextChanged.connect(self._on_template_changed)
        tpl_row.addWidget(self._template_combo, stretch=1)
        layout.addLayout(tpl_row)

        # ---- Sheet selector ----
        sheet_row = QHBoxLayout()
        sht_lbl = QLabel("Sheet:")
        sheet_row.addWidget(sht_lbl)
        sht_lbl.setFixedWidth(55)           # smoooooootthhhhh oooperrratorrrr
        self._sheet_combo = QComboBox()
        self._sheet_combo.setEditable(False)
        self._sheet_combo.currentTextChanged.connect(self._on_sheet_changed)
        sheet_row.addWidget(self._sheet_combo, stretch=1)
        layout.addLayout(sheet_row)

        # ---- Column mapping display ----
        col_grid = QGridLayout()
        self._col_edits: list[QLineEdit] = []
        for i, label in enumerate(_FIELD_LABELS):
            col_grid.addWidget(QLabel(f"{label}:"), i // 2, (i % 2) * 2)
            edit = QLineEdit()
            edit.setMaxLength(3)
            edit.setFixedWidth(50)
            edit.setToolTip(f"Excel column letter for '{label}'")
            col_grid.addWidget(edit, i // 2, (i % 2) * 2 + 1)
            self._col_edits.append(edit)
        layout.addLayout(col_grid)

        # ---- Manage profiles button ----
        manage_btn = QPushButton("Manage Template Profiles…")
        manage_btn.clicked.connect(self._on_manage_profiles)
        layout.addWidget(manage_btn)

        # Container to enable/disable template-specific widgets together
        self._template_widgets = [
            self._template_combo, self._sheet_combo, manage_btn,
            *self._col_edits,
        ]

        self._refresh_template_list()
        self._set_template_controls_enabled(False)

        return grp

    def _refresh_template_list(self) -> None:
        current = self._template_combo.currentText()
        self._template_combo.blockSignals(True)
        self._template_combo.clear()
        self._template_combo.addItems(self._cfg.available_templates())
        if current and self._template_combo.findText(current) >= 0:
            self._template_combo.setCurrentText(current)
        self._template_combo.blockSignals(False)
        self._on_template_changed(self._template_combo.currentText())

    def _on_mode_toggled(self, scratch_checked: bool) -> None:
        self._set_template_controls_enabled(not scratch_checked)

    def _set_template_controls_enabled(self, enabled: bool) -> None:
        for w in self._template_widgets:
            w.setEnabled(enabled)

    def _on_template_changed(self, name: str) -> None:
        """Refresh the sheet combo from the workbook, then load the saved sheet selection."""
        saved_sheet = self._cfg.get_selected_sheet(name)
        self._sheet_combo.blockSignals(True)
        self._sheet_combo.clear()
        self._sheet_combo.addItems(_read_sheet_names(name))
        idx = self._sheet_combo.findText(saved_sheet)
        if idx >= 0:
            self._sheet_combo.setCurrentIndex(idx)
        self._sheet_combo.blockSignals(False)
        self._on_sheet_changed(self._sheet_combo.currentText())

    def _on_sheet_changed(self, sheet: str) -> None:
        """Load the column mapping for the currently selected (template, sheet) pair."""
        name = self._template_combo.currentText()
        cols = self._cfg.get_column_mapping(name, sheet)
        for edit, col in zip(self._col_edits, cols):
            edit.setText(col)

    def _on_manage_profiles(self) -> None:
        dlg = _TemplateProfileDialog(self._cfg, self)
        dlg.exec()
        # Refresh sheet list and column display after dialog closes
        self._on_template_changed(self._template_combo.currentText())

    def apply_to_config(self) -> None:
        self._cfg.output_mode       = OUTPUT_MODE_SCRATCH if self._scratch_radio.isChecked() else OUTPUT_MODE_TEMPLATE
        self._cfg.selected_template = self._template_combo.currentText()
        selected_sheet              = self._sheet_combo.currentText()
        if self._cfg.selected_template:
            # Persist selected sheet
            self._cfg.template_selected_sheets[self._cfg.selected_template] = selected_sheet
            # Persist the currently displayed column mapping for this (template, sheet) pair
            if selected_sheet:
                cols = [e.text().strip().upper() for e in self._col_edits]
                self._cfg.template_profiles.setdefault(self._cfg.selected_template, {})[selected_sheet] = cols

    def load_from_config(self) -> None:
        is_template = self._cfg.output_mode == OUTPUT_MODE_TEMPLATE
        self._template_radio.setChecked(is_template)
        self._scratch_radio.setChecked(not is_template)
        self._set_template_controls_enabled(is_template)
        self._refresh_template_list()
        if self._cfg.selected_template:
            idx = self._template_combo.findText(self._cfg.selected_template)
            if idx >= 0:
                self._template_combo.setCurrentIndex(idx)
        self._on_template_changed(self._template_combo.currentText())

    def validation_errors(self) -> list[str]:
        if self._scratch_radio.isChecked():
            return []
        errors: list[str] = []
        if not self._template_combo.currentText():
            errors.append("No template selected for template-based output.")
            return errors
        if not self._sheet_combo.currentText():
            errors.append("No sheet selected. Add a profile for this template via 'Manage Template Profiles'.")
            return errors
        import openpyxl.utils as _oxl_utils
        for label, edit in zip(_FIELD_LABELS, self._col_edits):
            try:
                if _oxl_utils.column_index_from_string(edit.text().strip()) > 16384:  # XFD is the Excel max
                    raise ValueError
            except ValueError:
                errors.append(f"Column for '{label}' must be a valid Excel column letter (e.g. A, B, XFD).")
        return errors

    def get_column_mapping(self) -> list[str]:
        """Return the current column letters from the UI edits."""
        return [e.text().strip().upper() for e in self._col_edits]


# ---------------------------------------------------------------------------
# Main window
# ---------------------------------------------------------------------------

class MainWindow(QMainWindow):
    """Top-level application window."""

    def __init__(self) -> None:
        super().__init__()
        self._cfg    = AppConfig.load()
        self._worker: _PipelineWorker | None = None
        self._build_ui()
        self._load_config_into_ui()

    def _build_ui(self) -> None:
        self.setWindowTitle("CPAuto")
        if APP_ICON_PATH.exists():
            self.setWindowIcon(QIcon(str(APP_ICON_PATH)))
        self.resize(1000, 680)
        self.showMaximized()

        central = QWidget()
        self.setCentralWidget(central)
        root = QVBoxLayout(central)
        root.setContentsMargins(8, 8, 8, 8)
        root.setSpacing(6)

        root.addWidget(self._make_file_bar())

        # Left: settings (scrollable), Right: change type tree
        splitter = QSplitter(Qt.Orientation.Horizontal)

        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        self._settings_panel = _SettingsPanel(self._cfg)
        scroll.setWidget(self._settings_panel)
        splitter.addWidget(scroll)

        self._tree_panel = _ChangeTypeTree(self._cfg)
        splitter.addWidget(self._tree_panel)
        splitter.setSizes([380, 580])

        root.addWidget(splitter, stretch=1)
        root.addWidget(self._make_action_bar())

        self._status = QStatusBar()
        self.setStatusBar(self._status)

    def _make_file_bar(self) -> QWidget:
        bar    = QWidget()
        layout = QHBoxLayout(bar)
        layout.setContentsMargins(0, 0, 0, 0)

        layout.addWidget(QLabel("Document:"))
        self._file_edit = QLineEdit()
        self._file_edit.setPlaceholderText("Select a .docx file …")
        self._file_edit.setReadOnly(True)
        self._file_edit.mousePressEvent = lambda _: self._on_browse()
        layout.addWidget(self._file_edit, stretch=1)

        browse_btn = QPushButton("Browse…")
        browse_btn.clicked.connect(self._on_browse)
        layout.addWidget(browse_btn)

        return bar

    def _make_action_bar(self) -> QWidget:
        bar    = QWidget()
        layout = QHBoxLayout(bar)
        layout.setContentsMargins(0, 0, 0, 0)

        self._reset_btn = QPushButton("Reset to Default")
        self._reset_btn.clicked.connect(self._on_reset)

        layout.addWidget(self._reset_btn)
        layout.addStretch()

        self._save_btn = QPushButton("Save Config")
        self._save_btn.clicked.connect(self._on_save_config)

        self._run_btn = QPushButton("▶  Run")
        self._run_btn.setDefault(True)
        self._run_btn.clicked.connect(self._on_run)

        layout.addWidget(self._save_btn)
        layout.addWidget(self._run_btn)

        return bar

    def _on_browse(self) -> None:
        path, _ = QFileDialog.getOpenFileName(
            self, "Select .docx file", "", "Word Documents (*.docx)"
        )
        if path:
            self._file_edit.setText(path)

    def _on_run(self) -> None:
        docx = self._file_edit.text().strip()
        if not docx:
            QMessageBox.warning(self, "No File", "Please select a .docx file first.")
            return

        errors = self._collect_config_from_ui()
        if errors:
            QMessageBox.critical(self, "Validation Errors", "\n".join(errors))
            return

        self._run_btn.setEnabled(False)
        self._status.showMessage("Running pipeline …")

        self._worker = _PipelineWorker(docx, self._cfg)
        self._worker.log.connect(self._on_pipeline_log)
        self._worker.finished.connect(self._on_pipeline_finished)
        self._worker.start()

    def _on_save_config(self) -> None:
        errors = self._collect_config_from_ui()
        if errors:
            QMessageBox.critical(self, "Validation Errors", "\n".join(errors))
            return
        self._cfg.save()
        self._status.showMessage("Config saved.", 3000)

    def _on_reset(self) -> None:
        self._cfg.reset_to_default()
        self._load_config_into_ui()
        self._status.showMessage("Settings reset to defaults.", 3000)

    def _on_pipeline_log(self, msg: str) -> None:
        self._status.showMessage(msg)

    def _on_pipeline_finished(self, success: bool, msg: str) -> None:
        self._run_btn.setEnabled(True)
        if success:
            self._status.showMessage(msg, 8000)
            QMessageBox.information(self, "Done", msg)
        else:
            self._status.showMessage(msg, 8000)
            QMessageBox.critical(self, "Pipeline Error", msg)

    def _load_config_into_ui(self) -> None:
        self._settings_panel.load_from_config()
        self._tree_panel.load_from_config()

    def _collect_config_from_ui(self) -> list[str]:
        self._settings_panel.apply_to_config()
        self._tree_panel.apply_to_config()
        errors = self._settings_panel.validation_errors()
        if not self._cfg.selected_change_types:
            errors.append("At least one change type must be selected.")
        return errors


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------

def launch() -> None:
    """Create the QApplication, play the splash, then show the main window."""
    app = QApplication(sys.argv)
    app.setApplicationName("CPAuto")

    win = MainWindow()

    if SPLASH_VIDEO_PATH.exists():
        splash = _SplashScreen()
        splash.finished.connect(win.show)
        splash.show_and_play()
    else:
        win.show()

    sys.exit(app.exec())

if __name__ == "__main__":
    launch()
